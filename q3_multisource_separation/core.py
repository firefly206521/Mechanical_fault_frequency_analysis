"""Numerical core for Q3 multi-source sinusoid separation.

The module reuses the Q1 GLRT calibration and Q2 preprocessing/statistical
helpers without modifying either earlier question.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable

import numpy as np
import openpyxl

from . import SEED
from q2_harmonic_recovery.core import (
    analytic_signal,
    golden_minimize,
    wrap_phase,
)


SPLIT_MULTIPLIERS = np.asarray([0.10, 0.15, 0.25, 0.40, 0.65, 1.00, 1.50, 2.00])


# [AI-1] 辅助 Q3 GLRT 参数接口设计：每步 α_step=0.005 控制累积虚警
@dataclass(frozen=True)
class GLRTConfig:
    """Dependency-light copy of the current Q1 GLRT parameter interface."""

    f_min: float = 0.05
    f_max: float = 49.5
    p_fa: float = 0.005
    glrt_mc: int = 500
    random_seed: int = SEED
    threshold: float | None = None
    conditional_threshold: float | None = None


def q1_compatible_glrt_stat(x: np.ndarray, fs: float) -> tuple[np.ndarray, np.ndarray]:
    """GLRT 统计量：手动 double-detrend 替代 scipy.detrend，与 Q1 保持精确兼容。"""
    # [AI-1] 辅助手动线性去趋势实现，保证跨模块统计量数值一致
    index = np.arange(len(x), dtype=float)
    centered_index = index - np.mean(index)
    centered_x = x - np.mean(x)
    slope = float(np.dot(centered_index, centered_x) / np.dot(centered_index, centered_index))
    y = centered_x - slope * centered_index
    spec = np.fft.rfft(y)
    frequencies = np.fft.rfftfreq(len(y), 1.0 / fs)
    statistic = 2.0 * np.abs(spec) ** 2 / (len(y) * np.var(y) + np.finfo(float).eps)  # [AI-1] 辅助 epsilon 防止零方差
    return frequencies, statistic


@lru_cache(maxsize=16)
def q1_compatible_glrt_threshold(n: int, fs: float, cfg: GLRTConfig) -> float:
    rng = np.random.default_rng(cfg.random_seed)
    maxima = np.empty(cfg.glrt_mc)
    for index in range(cfg.glrt_mc):
        frequencies, statistic = q1_compatible_glrt_stat(rng.normal(0.0, 1.0, n), fs)
        mask = (frequencies >= cfg.f_min) & (frequencies <= cfg.f_max)
        maxima[index] = float(np.max(statistic[mask]))
    return float(np.quantile(maxima, 1.0 - cfg.p_fa))


def load_multi_source(path: Path) -> tuple[np.ndarray, np.ndarray, float]:
    """加载"多故障源"工作表：自动回退至第二工作表，校验时间严格递增。"""
    # [AI-1] 辅助多源工作表回退逻辑与递增校验
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "多故障源" in wb.sheetnames:
            ws = wb["多故障源"]
        elif len(wb.worksheets) >= 2:
            ws = wb.worksheets[1]
        else:
            raise ValueError("data.xlsx中找不到“多故障源”工作表，且不存在第二个工作表。")
        rows = [(row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True)]
    finally:
        wb.close()
    if not rows:
        raise ValueError("“多故障源”工作表没有数据行。")
    data = np.asarray(rows, dtype=float)
    if data.ndim != 2 or data.shape[1] < 2 or not np.all(np.isfinite(data[:, :2])):
        raise ValueError("“多故障源”工作表前两列必须是有限数值t和x(t)。")
    t, x = data[:, 0], data[:, 1]
    if len(t) < 3 or np.any(np.diff(t) <= 0):
        raise ValueError("时间列必须严格递增且至少包含3个样本。")
    fs = float(1.0 / np.median(np.diff(t)))
    return t, x, fs


def compute_bic(n: int, sse: float, parameter_count: int) -> float:
    """BIC 准则：零残差保护避免 log(0)。"""
    return float(n * math.log(max(sse / n, 1e-300)) + parameter_count * math.log(n))  # [AI-1] 辅助零残差 epsilon 保护


def _design_matrix(t: np.ndarray, frequencies: Iterable[float]) -> tuple[np.ndarray, float]:
    """构建多频设计矩阵：每频率一组 sin/cos 列 + offset 列，返回时间中心化矩阵。"""
    # [AI-1] 辅助向量化设计矩阵构建，支持任意数量频率分量
    frequencies = np.asarray(list(frequencies), dtype=float)
    center = float(np.mean(t))
    tc = t - center
    columns = []
    for frequency in frequencies:
        angle = 2.0 * np.pi * frequency * tc
        columns.extend([np.sin(angle), np.cos(angle)])
    columns.append(np.ones_like(t))
    return np.column_stack(columns), center


def design_diagnostics(design: np.ndarray, rcond: float = 1e-12) -> dict:
    """列归一化 SVD 诊断：条件数、秩、病态标志。"""
    norms = np.linalg.norm(design, axis=0)
    scaled = design / np.maximum(norms, np.finfo(float).eps)  # [AI-1] 辅助列归一化消除尺度对条件数影响
    singular = np.linalg.svd(scaled, compute_uv=False)
    threshold = rcond * singular[0]
    rank = int(np.sum(singular > threshold))
    condition = float(singular[0] / max(singular[-1], np.finfo(float).tiny))
    return {
        "condition_design": condition,
        "condition_normal": condition * condition,
        "smallest_singular_value": float(singular[-1]),
        "numerical_rank": rank,
        "column_count": design.shape[1],
        "ill_conditioned": bool(condition > 1e6 or rank < design.shape[1]),
        "warning_conditioned": bool(condition > 1e4),
    }


def multi_harmonic_fit(t: np.ndarray, y: np.ndarray, frequencies: Iterable[float]) -> dict:
    """多频联合最小二乘：对给定频率集做多谐波回归，返回各分量参数、BIC 和条件数诊断。"""
    # [AI-1] 辅助联合谐波回归，SVD 求解替代正规方程
    frequencies = np.sort(np.asarray(list(frequencies), dtype=float))
    design, center = _design_matrix(t, frequencies)
    beta, _, rank, _ = np.linalg.lstsq(design, y, rcond=1e-12)
    fitted = design @ beta
    residual = y - fitted
    diagnostics = design_diagnostics(design)
    diagnostics["lstsq_rank"] = int(rank)
    components = []
    for index, frequency in enumerate(frequencies):
        a, b = map(float, beta[2 * index:2 * index + 2])
        amplitude = float(np.hypot(a, b))
        phase_center = float(math.atan2(b, a))
        phase_origin = float(wrap_phase(phase_center - 2.0 * np.pi * frequency * center))
        component = a * design[:, 2 * index] + b * design[:, 2 * index + 1]
        components.append({
            "component_id": index + 1,
            "frequency_hz": float(frequency),
            "a_centered": a,
            "b_centered": b,
            "amplitude": amplitude,
            "phase_origin_rad": phase_origin,
            "waveform": component,
        })
    n = len(y)
    k = len(frequencies)
    parameter_count = 3 * k + 1
    sse = float(np.dot(residual, residual))
    bic = compute_bic(n, sse, parameter_count)
    return {
        "frequencies_hz": frequencies,
        "components": components,
        "offset": float(beta[-1]),
        "fit": fitted,
        "residual": residual,
        "sse": sse,
        "rmse": float(np.sqrt(sse / n)),
        "bic": bic,
        "parameter_count": parameter_count,
        "explained_variance": float(1.0 - np.var(residual) / np.var(y)),
        **diagnostics,
    }


def _multi_sse(t: np.ndarray, y: np.ndarray, frequencies: Iterable[float]) -> float:
    design, _ = _design_matrix(t, frequencies)
    beta, _, rank, _ = np.linalg.lstsq(design, y, rcond=1e-12)
    if rank < design.shape[1]:
        return float("inf")
    residual = y - design @ beta
    return float(np.dot(residual, residual))


def _sse_from_columns_cached(columns: list[np.ndarray], y: np.ndarray, rcond: float = 1e-12) -> float:
    """Return least-squares SSE from precomputed columns without stacking rows."""
    gram = np.empty((len(columns), len(columns)), dtype=float)
    rhs = np.empty(len(columns), dtype=float)
    for i, left in enumerate(columns):
        rhs[i] = float(np.dot(left, y))
        for j in range(i, len(columns)):
            value = float(np.dot(left, columns[j]))
            gram[i, j] = value
            gram[j, i] = value
    singular = np.linalg.svd(gram, compute_uv=False)
    if singular[-1] <= rcond * singular[0]:
        return float("inf")
    beta = np.linalg.solve(gram, rhs)
    yty = float(np.dot(y, y))
    return float(max(yty - float(np.dot(rhs, beta)), 0.0))


def _multi_sse_cached(t: np.ndarray, y: np.ndarray, frequencies: Iterable[float]) -> float:
    frequencies = np.asarray(list(frequencies), dtype=float)
    tc = t - float(np.mean(t))
    columns: list[np.ndarray] = []
    for frequency in frequencies:
        angle = 2.0 * np.pi * frequency * tc
        columns.extend([np.sin(angle), np.cos(angle)])
    columns.append(np.ones_like(t))
    return _sse_from_columns_cached(columns, y)


def refine_joint_frequencies(
    t: np.ndarray,
    y: np.ndarray,
    seeds: Iterable[float],
    half_width: float | Iterable[float] = 0.006,
    maxiter: int = 45,
    fit_backend: str = "dense",
) -> dict:
    """坐标下降联合精修：逐维黄金分割搜索，多轮迭代直至频率收敛或 maxiter 耗尽。"""
    # [AI-1] 辅助 cached backend 实现，通过预计算 Gram 矩阵加速联合精修
    if fit_backend not in {"dense", "cached"}:
        raise ValueError("fit_backend must be 'dense' or 'cached'")
    seeds = np.sort(np.asarray(list(seeds), dtype=float))
    if len(seeds) == 0:
        return multi_harmonic_fit(t, y, [])
    widths = np.full(len(seeds), float(half_width)) if np.isscalar(half_width) else np.asarray(list(half_width), float)
    base_bounds = [(max(0.001, f - w), f + w) for f, w in zip(seeds, widths)]
    frequencies = seeds.copy()
    previous = frequencies.copy()
    tc = t - float(np.mean(t))
    columns: list[np.ndarray] = []
    for frequency in frequencies:
        angle = 2.0 * np.pi * frequency * tc
        columns.extend([np.sin(angle), np.cos(angle)])
    offset_column = np.ones_like(t)
    sweeps = min(4, max(2, maxiter // 8))
    for _ in range(sweeps):
        for index in range(len(frequencies)):
            lo, hi = base_bounds[index]
            if index > 0:
                lo = max(lo, frequencies[index - 1] + 1e-7)
            if index + 1 < len(frequencies):
                hi = min(hi, frequencies[index + 1] - 1e-7)
            if hi <= lo:
                continue
            def objective(value: float) -> float:
                angle = 2.0 * np.pi * value * tc
                candidate_columns = columns.copy()
                candidate_columns[2 * index] = np.sin(angle)
                candidate_columns[2 * index + 1] = np.cos(angle)
                if fit_backend == "cached":
                    return _sse_from_columns_cached([*candidate_columns, offset_column], y)
                design = np.column_stack([*candidate_columns, offset_column])
                beta, _, rank, _ = np.linalg.lstsq(design, y, rcond=1e-12)
                if rank < design.shape[1]:
                    return float("inf")
                residual = y - design @ beta
                return float(np.dot(residual, residual))
            frequencies[index] = golden_minimize(objective, lo, hi, iterations=18)
            angle = 2.0 * np.pi * frequencies[index] * tc
            columns[2 * index] = np.sin(angle)
            columns[2 * index + 1] = np.cos(angle)
        if np.max(np.abs(frequencies - previous)) < 1e-9:
            break
        previous = frequencies.copy()
    fit = multi_harmonic_fit(t, y, frequencies)
    fit["optimizer_success"] = True
    fit["optimizer_message"] = "bounded coordinate golden-section refinement"
    return fit


def glrt_scan(residual: np.ndarray, fs: float, cfg: GLRTConfig, fitted_parameter_count: int) -> dict:
    """残差 GLRT 扫描：对当前残差计算周期图，定位最强剩余峰，用于 SIC 下一轮候选。"""
    # [AI-1] 辅助 SIC 残差重扫描管线，支持条件门限回退
    frequencies, statistic = q1_compatible_glrt_stat(residual, fs)
    mask = (frequencies >= cfg.f_min) & (frequencies <= cfg.f_max)
    corrected = statistic
    indexes = np.where(mask)[0]
    local = indexes[(corrected[indexes] >= np.roll(corrected, 1)[indexes]) & (corrected[indexes] > np.roll(corrected, -1)[indexes])]
    if len(local) == 0:
        local = indexes
    order = local[np.argsort(corrected[local])[::-1]]
    threshold = cfg.threshold if cfg.threshold is not None else q1_compatible_glrt_threshold(len(residual), fs, cfg)
    return {
        "frequencies": frequencies,
        "statistic": corrected,
        "peak_indexes": order,
        "peak_frequency_hz": float(frequencies[order[0]]),
        "peak_score": float(corrected[order[0]]),
        "threshold": float(threshold),
        "detected": bool(corrected[order[0]] >= threshold),
    }


def _candidate_seed_sets(current: np.ndarray, new_peak: float, duration: float) -> list[tuple[str, np.ndarray]]:
    """生成候选种子集：残差峰添加 + 近频拆分候选。"""
    # [AI-1] 辅助近频拆分网格设计（SPLIT_MULTIPLIERS 控制分离间隔）
    candidates: list[tuple[str, np.ndarray]] = []
    if len(current) == 0 or np.min(np.abs(current - new_peak)) > 1e-7:
        candidates.append(("residual_peak", np.sort(np.r_[current, new_peak])))
    if len(current) and np.min(np.abs(current - new_peak)) < 0.02:
        for index, center in enumerate(current):
            if abs(center - new_peak) >= 0.02:
                continue
            for multiplier in SPLIT_MULTIPLIERS:
                separation = float(multiplier / duration)
                pair = np.asarray([center - separation / 2.0, center + separation / 2.0])
                seed = np.sort(np.r_[np.delete(current, index), pair])
                candidates.append((f"split_{multiplier:.2f}_over_T", seed))
    return candidates


def detect_multitone(
    t: np.ndarray,
    y: np.ndarray,
    fs: float,
    cfg: GLRTConfig,
    max_components: int = 10,
    bic_delta: float = 10.0,
    fit_backend: str = "dense",
) -> tuple[dict, list[dict]]:
    """SIC+BIC 自动定阶：循环 {GLRT扫描残差→联合精修→BIC判据(Δ≥10)→接受/停止}。"""
    # [AI-1] 辅助 SIC 主循环编排与 BIC Δ≥10 接受判据
    if fit_backend not in {"dense", "cached"}:
        raise ValueError("fit_backend must be 'dense' or 'cached'")
    current = multi_harmonic_fit(t, y, [])
    history: list[dict] = []
    duration = float(t[-1] - t[0])
    for iteration in range(1, max_components + 1):
        scan = glrt_scan(current["residual"], fs, cfg, current["parameter_count"])
        record = {
            "iteration": iteration,
            "components_before": len(current["frequencies_hz"]),
            "residual_peak_hz": scan["peak_frequency_hz"],
            "glrt_score": scan["peak_score"],
            "glrt_threshold": scan["threshold"],
            "glrt_passed": scan["detected"],
            "bic_before": current["bic"],
        }
        if not scan["detected"]:
            record.update({"accepted": False, "stop_reason": "GLRT below threshold"})
            history.append(record)
            break
        seed_sets = _candidate_seed_sets(current["frequencies_hz"], scan["peak_frequency_hz"], duration)
        if len(current["frequencies_hz"]):
            nearest_index = int(np.argmin(np.abs(current["frequencies_hz"] - scan["peak_frequency_hz"])))
            nearest = float(current["frequencies_hz"][nearest_index])
            if abs(nearest - scan["peak_frequency_hz"]) < 0.02:
                local_center = 0.5 * (nearest + scan["peak_frequency_hz"])
                local = close_pair_resolver(t, y, center_hz=local_center, conditional_threshold=cfg.conditional_threshold)
                if local["estimated_k"] == 2:
                    local_seed = np.sort(np.r_[np.delete(current["frequencies_hz"], nearest_index), local["frequencies_hz"]])
                    seed_sets.append(("local_close_pair_resolver", local_seed))
        quick = []
        for origin, seeds in seed_sets:
            if len(seeds) > max_components or np.min(seeds) <= cfg.f_min or np.max(seeds) >= cfg.f_max:
                continue
            trial = multi_harmonic_fit(t, y, seeds)
            quick.append((trial["bic"], origin, seeds))
        quick.sort(key=lambda item: item[0])
        refined = []
        # [AI-1] 辅助快速 BIC 预筛：仅最优种子进入精修，节约计算
        for _, origin, seeds in quick[:1]:
            is_wide_residual_peak = origin == "residual_peak" and (
                len(current["frequencies_hz"]) == 0
                or np.min(np.abs(current["frequencies_hz"] - scan["peak_frequency_hz"])) >= 0.02
            )
            if is_wide_residual_peak:
                candidate = seeds.copy()
                new_index = int(np.argmin(np.abs(candidate - scan["peak_frequency_hz"])))
                lo = max(cfg.f_min, candidate[new_index] - max(3.0 / duration, 0.003))
                hi = min(cfg.f_max, candidate[new_index] + max(3.0 / duration, 0.003))
                sse_func = _multi_sse_cached if fit_backend == "cached" else _multi_sse
                candidate[new_index] = golden_minimize(
                    lambda value: sse_func(t, y, np.sort(np.r_[candidate[:new_index], value, candidate[new_index + 1:]])),
                    lo,
                    hi,
                    iterations=20,
                )
                trial = multi_harmonic_fit(t, y, np.sort(candidate))
            else:
                widths = np.full(len(seeds), max(3.0 / duration, 0.003))
                trial = refine_joint_frequencies(t, y, seeds, widths, fit_backend=fit_backend)
            refined.append((trial["bic"], origin, trial))
        if not refined:
            record.update({"accepted": False, "stop_reason": "no valid candidate"})
            history.append(record)
            break
        _, origin, best = min(refined, key=lambda item: item[0])
        improvement = current["bic"] - best["bic"]
        accepted = bool(improvement >= bic_delta and not best["ill_conditioned"])
        record.update({
            "candidate_origin": origin,
            "candidate_frequencies_hz": ";".join(f"{f:.12g}" for f in best["frequencies_hz"]),
            "bic_after": best["bic"],
            "bic_improvement": improvement,
            "condition_design": best["condition_design"],
            "accepted": accepted,
            "stop_reason": "accepted" if accepted else ("ill-conditioned" if best["ill_conditioned"] else "BIC improvement below 10"),
        })
        history.append(record)
        if not accepted:
            break
        current = best
    return current, history


def component_table(fit: dict) -> list[dict]:
    """分量 SNR 表：每分量的信号功率与残差噪声比。"""
    # [AI-1] 辅助分量级 SNR 估计与条件数标注
    noise_power = float(np.mean(fit["residual"] ** 2))
    rows = []
    for component in fit["components"]:
        signal_power = component["amplitude"] ** 2 / 2.0
        rows.append({
            "component_id": component["component_id"],
            "frequency_hz": component["frequency_hz"],
            "amplitude": component["amplitude"],
            "phase_origin_rad": component["phase_origin_rad"],
            "component_snr_db": 10.0 * math.log10(signal_power / noise_power),
            "condition_design": fit["condition_design"],
            "condition_normal": fit["condition_normal"],
            "ill_conditioned": fit["ill_conditioned"],
        })
    return rows


def segment_stability(t: np.ndarray, y: np.ndarray, frequencies: np.ndarray, seconds: float = 50.0) -> list[dict]:
    """分段稳定性分析：各段独立/公共频率拟合，评估频率一致性。"""
    # [AI-1] 辅助分段公共频率与独立频率偏差统计
    fs = 1.0 / np.median(np.diff(t))
    size = int(round(seconds * fs))
    rows = []
    for segment_id, start in enumerate(range(0, len(y) - size + 1, size), 1):
        end = start + size
        st, sy = t[start:end], y[start:end]
        common = multi_harmonic_fit(st, sy, frequencies)
        independent = refine_joint_frequencies(st, sy, frequencies, half_width=0.01, maxiter=25)
        for index, (base, refined, component) in enumerate(zip(frequencies, independent["frequencies_hz"], common["components"]), 1):
            rows.append({
                "segment_id": segment_id,
                "start_s": float(st[0]),
                "end_s": float(st[-1]),
                "component_id": index,
                "common_frequency_hz": float(base),
                "independent_frequency_hz": float(refined),
                "frequency_deviation_hz": float(refined - base),
                "amplitude": component["amplitude"],
                "phase_origin_rad": component["phase_origin_rad"],
                "residual_std": float(np.std(common["residual"])),
            })
    return rows


def baseband_series(t: np.ndarray, x: np.ndarray, center_hz: float = 13.5, target_fs: float = 1.0) -> tuple[np.ndarray, np.ndarray, float]:
    """基带降采样：Hilbert→复混频→低通→抽取，将窄带信号搬至零频。"""
    # [AI-1] 辅助复混频基带变换与 FFT 低通抽取
    analytic = analytic_signal(x)
    mixed = analytic * np.exp(-1j * 2.0 * np.pi * center_hz * t)
    fs = 1.0 / np.median(np.diff(t))
    freq = np.fft.fftfreq(len(t), 1.0 / fs)
    af = np.abs(freq)
    weight = np.zeros_like(freq)
    weight[af <= 0.03] = 1.0
    taper = (af > 0.03) & (af < 0.05)
    weight[taper] = 0.5 * (1.0 + np.cos(np.pi * (af[taper] - 0.03) / 0.02))
    filtered = np.fft.ifft(np.fft.fft(mixed) * weight)
    factor = max(1, int(round(fs / target_fs)))
    return t[::factor], filtered[::factor], fs / factor


def _complex_fit(t: np.ndarray, z: np.ndarray, offsets: Iterable[float]) -> dict:
    """复指数联合拟合：多频偏移→复设计矩阵→最小二乘→BIC 评估。"""
    # [AI-1] 辅助复设计矩阵 SVD 诊断与 BIC 计算
    offsets = np.sort(np.asarray(list(offsets), dtype=float))
    columns = [np.exp(1j * 2.0 * np.pi * offset * t) for offset in offsets]
    columns.append(np.ones_like(t, dtype=complex))
    design = np.column_stack(columns)
    beta, _, rank, _ = np.linalg.lstsq(design, z, rcond=1e-12)
    residual = z - design @ beta
    diag = design_diagnostics(design)
    n_real = 2 * len(z)
    parameter_count = 3 * len(offsets) + 2
    sse = float(np.sum(np.abs(residual) ** 2))
    bic = compute_bic(n_real, sse, parameter_count)
    return {"offsets": offsets, "beta": beta, "residual": residual, "sse": sse, "bic": bic, "rank": int(rank), **diag}


def _complex_sse(t: np.ndarray, z: np.ndarray, offsets: Iterable[float]) -> float:
    """复指数拟合 SSE：构建设计矩阵→最小二乘→残差平方和。"""
    # [AI-1] 辅助复指数设计矩阵与 SSE 快速计算
    offsets = np.asarray(list(offsets), dtype=float)
    design = np.column_stack([*[np.exp(1j * 2.0 * np.pi * offset * t) for offset in offsets], np.ones_like(t, dtype=complex)])
    beta, _, rank, _ = np.linalg.lstsq(design, z, rcond=1e-12)
    if rank < design.shape[1]:
        return float("inf")
    residual = z - design @ beta
    return float(np.sum(np.abs(residual) ** 2))


def _baseband_periodogram(z: np.ndarray, fs: float, step_hz: float, limit_hz: float = 0.012) -> tuple[np.ndarray, np.ndarray]:
    """Fine local periodogram computed by zero-padded complex FFT."""
    # [AI-1] 辅助零填充复 FFT 高分辨率局部周期图
    nfft = max(len(z), int(math.ceil(fs / step_hz)))
    centered = z - np.mean(z)
    frequencies = np.fft.fftfreq(nfft, 1.0 / fs)
    power = np.abs(np.fft.fft(centered, n=nfft)) ** 2
    mask = np.abs(frequencies) <= limit_hz
    order = np.argsort(frequencies[mask])
    return frequencies[mask][order], power[mask][order]


def _refine_complex_pair(t: np.ndarray, z: np.ndarray, seed: np.ndarray, rounds: int = 3) -> dict:
    values = np.sort(np.asarray(seed, dtype=float))
    for _ in range(rounds):
        left_hi = min(values[1] - 1e-7, 0.012)
        if left_hi > -0.012:
            values[0] = golden_minimize(
                lambda value: _complex_sse(t, z, [value, values[1]]),
                -0.012,
                left_hi,
                iterations=18,
            )
        right_lo = max(values[0] + 1e-7, -0.012)
        if right_lo < 0.012:
            values[1] = golden_minimize(
                lambda value: _complex_sse(t, z, [values[0], value]),
                right_lo,
                0.012,
                iterations=18,
            )
    return _complex_fit(t, z, values)


def _refine_single_full(t: np.ndarray, x: np.ndarray, seed_hz: float, width_hz: float = 0.00025) -> dict:
    """Refine the one-tone null on the full real-valued record."""
    # [AI-1] 辅助全记录单频零假设黄金分割精修
    frequency = golden_minimize(
        lambda value: _multi_sse(t, x, [value]),
        max(0.001, seed_hz - width_hz),
        seed_hz + width_hz,
        iterations=18,
    )
    return multi_harmonic_fit(t, x, [frequency])


def _pattern_refine_pair_full(
    t: np.ndarray,
    x: np.ndarray,
    seed_hz: np.ndarray,
    initial_step_hz: float = 0.0002,
    max_iterations: int = 8,
) -> dict:
    """Two-dimensional variable-projection search in center/separation space."""
    # [AI-1] 辅助二维中心/分离度交替下降搜索与 SSE 缓存
    seed = np.sort(np.asarray(seed_hz, float))
    center = float(np.mean(seed))
    separation = float(seed[1] - seed[0])
    center_step = initial_step_hz
    separation_step = 2.0 * initial_step_hz
    cache: dict[tuple[float, float], float] = {}

    def objective(candidate_center: float, candidate_separation: float) -> float:
        if not 1e-7 <= candidate_separation <= 0.024:
            return float("inf")
        frequencies = (candidate_center - candidate_separation / 2.0, candidate_center + candidate_separation / 2.0)
        if frequencies[0] <= 0.001:
            return float("inf")
        key = (round(candidate_center, 12), round(candidate_separation, 12))
        if key not in cache:
            cache[key] = _multi_sse(t, x, frequencies)
        return cache[key]

    best = objective(center, separation)
    for _ in range(max_iterations):
        candidates = []
        for dc, ds in ((-1, 0), (1, 0), (0, -1), (0, 1)):
            trial_center = center + dc * center_step
            trial_separation = separation + ds * separation_step
            candidates.append((objective(trial_center, trial_separation), trial_center, trial_separation))
        trial_sse, trial_center, trial_separation = min(candidates, key=lambda row: row[0])
        if trial_sse + 1e-12 < best:
            best, center, separation = trial_sse, trial_center, trial_separation
        center_step *= 0.5
        separation_step *= 0.5
        if max(center_step, separation_step) < 1e-8:
            break
    frequencies = np.asarray([center - separation / 2.0, center + separation / 2.0])
    fit = multi_harmonic_fit(t, x, frequencies)
    fit["optimizer_message"] = "multi-start full-record center/separation pattern search"
    return fit


def _frequency_uncertainty(t: np.ndarray, x: np.ndarray, fit: dict) -> dict:
    """Finite-difference Hessian uncertainty for a fitted frequency pair."""
    # [AI-1] 辅助有限差分 Hessian 与频率标准误 95% CI 估计
    frequencies = np.asarray(fit["frequencies_hz"], float)
    if len(frequencies) != 2:
        return {"frequency_ci_reliable": False}
    duration = float(t[-1] - t[0])
    h = max(1e-7, 1.0 / (200.0 * duration))
    base = _multi_sse(t, x, frequencies)
    hessian = np.empty((2, 2), float)
    for index in range(2):
        plus = frequencies.copy(); plus[index] += h
        minus = frequencies.copy(); minus[index] -= h
        hessian[index, index] = (_multi_sse(t, x, plus) - 2.0 * base + _multi_sse(t, x, minus)) / h ** 2
    pp = _multi_sse(t, x, frequencies + np.asarray([h, h]))
    pm = _multi_sse(t, x, frequencies + np.asarray([h, -h]))
    mp = _multi_sse(t, x, frequencies + np.asarray([-h, h]))
    mm = _multi_sse(t, x, frequencies - np.asarray([h, h]))
    hessian[0, 1] = hessian[1, 0] = (pp - pm - mp + mm) / (4.0 * h ** 2)
    residual_variance = fit["sse"] / max(len(t) - fit["parameter_count"], 1)
    reliable = bool(np.all(np.isfinite(hessian)) and np.min(np.linalg.eigvalsh(hessian)) > 0.0)
    if not reliable:
        return {"frequency_ci_reliable": False}
    covariance = 2.0 * residual_variance * np.linalg.inv(hessian)
    standard_errors = np.sqrt(np.maximum(np.diag(covariance), 0.0))
    if not np.all(np.isfinite(standard_errors)) or np.max(standard_errors) > 0.012:
        return {"frequency_ci_reliable": False}
    return {
        "frequency_ci_reliable": True,
        "frequency_1_se_hz": float(standard_errors[0]),
        "frequency_2_se_hz": float(standard_errors[1]),
        "frequency_1_ci95_low_hz": float(frequencies[0] - 1.96 * standard_errors[0]),
        "frequency_1_ci95_high_hz": float(frequencies[0] + 1.96 * standard_errors[0]),
        "frequency_2_ci95_low_hz": float(frequencies[1] - 1.96 * standard_errors[1]),
        "frequency_2_ci95_high_hz": float(frequencies[1] + 1.96 * standard_errors[1]),
    }


@lru_cache(maxsize=16)
def conditional_glrt_threshold(
    n: int,
    fs: float,
    local_grid_step_hz: float,
    monte_carlo_runs: int = 500,
    alpha: float = 0.01,
    random_seed: int = SEED + 31,
) -> float:
    """Monte Carlo threshold for a searched second tone under a one-tone null."""
    # [AI-1] 辅助条件 GLRT MC 门限：单频零假设下的 F 统计量 99% 分位
    rng = np.random.default_rng(random_seed)
    time_axis = np.arange(n, dtype=float) / fs
    statistics = np.empty(monte_carlo_runs, float)
    for replicate in range(monte_carlo_runs):
        noise = (rng.normal(size=n) + 1j * rng.normal(size=n)) / math.sqrt(2.0)
        one = _complex_fit(time_axis, noise, [0.0])
        grid, power = _baseband_periodogram(one["residual"], fs, local_grid_step_hz)
        allowed = np.abs(grid) >= max(4.0 * local_grid_step_hz, 0.00015)
        candidate = float(grid[int(np.argmax(np.where(allowed, power, -np.inf)))])
        two = _complex_fit(time_axis, noise, [0.0, candidate])
        numerator = max(one["sse"] - two["sse"], 0.0) / 3.0
        denominator = two["sse"] / max(2 * n - 8, 1)
        statistics[replicate] = numerator / max(denominator, np.finfo(float).eps)
    return float(np.quantile(statistics, 1.0 - alpha))


def close_pair_resolver(
    t: np.ndarray,
    x: np.ndarray,
    center_hz: float = 13.5,
    local_grid_step_hz: float = 0.000025,
    conditional_threshold: float | None = None,
) -> dict:
    """Resolve a local two-tone cluster using peel, rescan, and joint refit.

    The center only defines the heterodyne passband. Candidate frequencies are
    estimated independently and are not constrained to be symmetric around it.
    """
    # [AI-1] 辅助 SIC 剥离+重扫描+联合拟合流程，含多候选策略评估
    tb, z, fsb = baseband_series(t, x, center_hz)
    grid, original_power = _baseband_periodogram(z, fsb, local_grid_step_hz)
    coarse_one = float(grid[int(np.argmax(original_power))])
    refine_width = max(4.0 * local_grid_step_hz, 0.0002)
    objective1 = lambda value: _complex_sse(tb, z, [float(value)])
    one_frequency = golden_minimize(
        objective1,
        max(-0.012, coarse_one - refine_width),
        min(0.012, coarse_one + refine_width),
        iterations=22,
    )
    one = _complex_fit(tb, z, [one_frequency])

    # Strong-component peeling exposes a weak asymmetric neighbour. The final
    # answer is still obtained from a joint two-tone fit, avoiding error
    # propagation from sequential subtraction.
    residual_grid, residual_power = _baseband_periodogram(one["residual"], fsb, local_grid_step_hz)
    exclusion = max(4.0 * local_grid_step_hz, 0.00015)
    allowed = np.abs(residual_grid - one_frequency) >= exclusion
    if np.any(allowed):
        residual_index = int(np.argmax(np.where(allowed, residual_power, -np.inf)))
        residual_frequency = float(residual_grid[residual_index])
        residual_peak_ratio = float(residual_power[residual_index] / max(np.median(residual_power[allowed]), 1e-30))
    else:
        residual_frequency = float(one_frequency)
        residual_peak_ratio = 0.0

    duration = float(t[-1] - t[0])
    candidates: list[tuple[str, np.ndarray]] = [("peel_residual_rescan", np.sort([one_frequency, residual_frequency]))]
    for multiplier in SPLIT_MULTIPLIERS:
        separation = float(multiplier / duration)
        candidates.append((f"adaptive_split_{multiplier:.2f}_over_T", np.asarray([
            one_frequency - separation / 2.0,
            one_frequency + separation / 2.0,
        ])))

    # Add pairs formed by the strongest distinct local spectral maxima.
    maxima = np.where((original_power[1:-1] > original_power[:-2]) & (original_power[1:-1] >= original_power[2:]))[0] + 1
    maxima = maxima[np.argsort(original_power[maxima])[::-1]]
    local_peaks: list[float] = []
    for index in maxima:
        value = float(grid[index])
        if all(abs(value - existing) >= exclusion for existing in local_peaks):
            local_peaks.append(value)
        if len(local_peaks) >= 3:
            break
    for value in local_peaks:
        if abs(value - one_frequency) >= exclusion:
            candidates.append(("fine_grid_secondary_peak", np.sort([one_frequency, value])))

    valid = [(origin, np.sort(seed)) for origin, seed in candidates if seed[0] >= -0.012 and seed[1] <= 0.012 and seed[1] - seed[0] >= 1e-7]
    quick = sorted((_complex_fit(tb, z, seed)["bic"], origin, seed) for origin, seed in valid)
    refined = []
    for _, origin, seed in quick[:3]:
        trial = _refine_complex_pair(tb, z, seed)
        refined.append((trial["bic"], origin, trial))
    _, candidate_origin, two = min(refined, key=lambda item: item[0])

    one_full_candidates = [
        _refine_single_full(t, x, center_hz + one["offsets"][0]),
        _refine_single_full(t, x, center_hz),
    ]
    one_full = min(one_full_candidates, key=lambda fit: fit["sse"])
    full_candidates = []
    for _, origin, trial in refined[:1]:
        seed_hz = center_hz + trial["offsets"]
        optimized = _pattern_refine_pair_full(
            t,
            x,
            seed_hz,
            initial_step_hz=max(20.0 * local_grid_step_hz, 0.0005),
        )
        full_candidates.append((optimized["bic"], origin, optimized))
    _, candidate_origin, full_fit = min(full_candidates, key=lambda item: item[0])
    frequencies = full_fit["frequencies_hz"]
    improvement = one_full["bic"] - full_fit["bic"]
    conditional_statistic = (
        max(one_full["sse"] - full_fit["sse"], 0.0) / 3.0
    ) / max(full_fit["sse"] / max(len(t) - full_fit["parameter_count"], 1), np.finfo(float).eps)
    if conditional_threshold is None:
        conditional_threshold = conditional_glrt_threshold(len(tb), fsb, local_grid_step_hz)
    accepted = bool(
        improvement >= 10.0
        and conditional_statistic >= conditional_threshold
        and full_fit["numerical_rank"] == full_fit["column_count"]
        and not full_fit["ill_conditioned"]
    )
    uncertainty = _frequency_uncertainty(t, x, full_fit) if accepted else {"frequency_ci_reliable": False}
    return {
        "estimated_k": 2 if accepted else 1,
        "frequencies_hz": frequencies if accepted else np.asarray([center_hz + one["offsets"][0]]),
        "bic_one": one["bic"],
        "bic_two": two["bic"],
        "bic_improvement": improvement,
        "condition_design": full_fit["condition_design"],
        "condition_normal": full_fit["condition_normal"],
        "smallest_singular_value": full_fit["smallest_singular_value"],
        "numerical_rank": full_fit["numerical_rank"],
        "column_count": full_fit["column_count"],
        "ill_conditioned": full_fit["ill_conditioned"],
        "baseband_fs_hz": fsb,
        "candidate_origin": candidate_origin,
        "residual_peak_ratio": residual_peak_ratio,
        "conditional_glrt_statistic": conditional_statistic,
        "conditional_glrt_threshold": conditional_threshold,
        "local_grid_step_hz": local_grid_step_hz,
        **uncertainty,
    }


def music_close_pair(
    t: np.ndarray,
    x: np.ndarray,
    center_hz: float = 13.5,
    grid_step_hz: float = 0.000025,
    order: int = 80,
) -> dict:
    """MUSIC 近频对照：基带降采样→协方差→MDL 定阶→伪谱峰值提取。"""
    # [AI-1] 辅助 MUSIC MDL 定阶与伪谱峰值搜索
    tb, z, fsb = baseband_series(t, x, center_hz)
    order = min(order, len(z) // 2)
    trajectory = np.lib.stride_tricks.sliding_window_view(z, order).T
    snapshots = trajectory.shape[1]
    covariance = trajectory @ trajectory.conj().T / snapshots
    eigenvalues, eigenvectors = np.linalg.eigh(covariance)
    eigenvalues = np.maximum(eigenvalues.real, 1e-300)
    mdl = []
    for k in range(0, min(4, order - 1) + 1):
        noise_values = eigenvalues[:order - k]
        ratio = np.exp(np.mean(np.log(noise_values))) / np.mean(noise_values)
        value = -snapshots * (order - k) * math.log(max(ratio, 1e-300)) + 0.5 * k * (2 * order - k) * math.log(snapshots)
        mdl.append(value)
    estimated_k = int(np.argmin(mdl))
    grid = np.arange(-0.0075, 0.0075 + grid_step_hz / 2, grid_step_hz)
    if estimated_k < 1:
        return {"estimated_k": 0, "frequencies_hz": np.asarray([]), "grid_step_hz": grid_step_hz, "mdl": mdl}
    signal_space = eigenvectors[:, -estimated_k:]
    steering = np.exp(1j * 2.0 * np.pi * np.arange(order)[:, None] * (1.0 / fsb) * grid[None, :])
    denominator = order - np.sum(np.abs(signal_space.conj().T @ steering) ** 2, axis=0)
    pseudo = 1.0 / np.maximum(denominator.real, 1e-15)
    peaks = np.where((pseudo[1:-1] > pseudo[:-2]) & (pseudo[1:-1] >= pseudo[2:]))[0] + 1
    peaks = peaks[np.argsort(pseudo[peaks])[::-1]] if len(peaks) else np.asarray([int(np.argmax(pseudo))])
    selected = np.sort(peaks[:min(estimated_k, 2)])
    return {
        "estimated_k": int(len(selected)),
        "frequencies_hz": center_hz + grid[selected],
        "grid_step_hz": grid_step_hz,
        "mdl": mdl,
        "grid_hz": center_hz + grid,
        "pseudospectrum": pseudo,
    }
